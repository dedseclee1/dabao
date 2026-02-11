# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
import os
import datetime
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import warnings
import win32com.client as win32  # <-- 新增导入
import pythoncom  # <-- 新增导入


# --- 辅助函数：标黄Excel行 (无修改) ---
def highlight_rows_in_excel(excel_path, wos_to_highlight_set):
    """
    打开指定的Excel文件，并标黄 'wos_to_highlight_set' 中包含的 (工单单号, 工单单别) 对应的行。
    """
    log_messages = []
    if not wos_to_highlight_set:
        return True, "INFO: (highlight) 没有需要标黄的工单。"

    try:
        # 定义黄色填充
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

        log_messages.append(f"INFO: (highlight) 正在加载 '{os.path.basename(excel_path)}' 以进行标黄...")
        warnings.simplefilter("ignore", category=UserWarning)
        workbook = load_workbook(excel_path)
        warnings.simplefilter("default", category=UserWarning)

        sheet = workbook.active
        log_messages.append(f"INFO: (highlight) 目标工作表: '{sheet.title}'")

        header_row = sheet[1]
        col_indices = {cell.value: cell.column for cell in header_row}

        if '工单单号' not in col_indices or '工单单别' not in col_indices:
            msg = f"错误: (highlight) 在 '{sheet.title}' 的表头中未找到 '工单单号' 或 '工单单别' 列。无法标黄。"
            log_messages.append(msg)
            return False, "\n".join(log_messages)

        wo_num_col_letter = get_column_letter(col_indices['工单单号'])
        wo_type_col_letter = get_column_letter(col_indices['工单单别'])

        highlighted_count = 0
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            wo_num = sheet[f"{wo_num_col_letter}{row[0].row}"].value
            wo_type = sheet[f"{wo_type_col_letter}{row[0].row}"].value

            if wo_num is None or wo_type is None:
                continue

            key = (str(wo_num).strip(), str(wo_type).strip())

            if key in wos_to_highlight_set:
                for cell in row:
                    cell.fill = yellow_fill
                highlighted_count += 1

        if highlighted_count > 0:
            log_messages.append(f"INFO: (highlight) 匹配到 {highlighted_count} 行待标黄。")
            workbook.save(excel_path)
            log_messages.append(f"成功: (highlight) '{os.path.basename(excel_path)}' 已更新并保存。")
        else:
            log_messages.append(
                f"INFO: (highlight) 待导入的工单在 '{os.path.basename(excel_path)}' 中未找到匹配行，无需标黄。")

        return True, "\n".join(log_messages)

    except FileNotFoundError:
        msg = f"错误: (highlight) 无法找到文件进行标黄: {excel_path}"
        log_messages.append(msg)
        return False, "\n".join(log_messages)
    except Exception as e:
        msg = f"错误: (highlight) 标黄文件 '{os.path.basename(excel_path)}' 时发生意外失败 - {e}\n{traceback.format_exc()}"
        log_messages.append(msg)
        return False, "\n".join(log_messages)


# --- [新增 V4.8] 辅助函数：在DataFrame中查找表头行 ---
def find_header_row_in_baseline(df_no_header):
    """
    在无表头的DataFrame (前10行) 中查找包含关键列的行。
    返回行索引 (0-based)，如果未找到则返回 -1。
    """
    for idx, row in df_no_header.head(10).iterrows():  # 仅检查前10行
        try:
            row_values_cleaned = [str(v).strip() for v in row.values]
            row_values_upper = [v.upper() for v in row_values_cleaned]

            # 检查关键列
            has_wo_num = '工单单号' in row_values_cleaned
            has_wo_type = '工单单别' in row_values_cleaned

            has_date = False
            for val_upper in row_values_upper:
                if 'MOCTA' in val_upper and 'CREATE' in val_upper and 'DATE' in val_upper:
                    has_date = True
                    break

            # 必须同时包含这三组
            if has_wo_num and has_wo_type and has_date:
                return idx  # 找到了! 这就是表头行的索引 (0-based)
        except Exception:
            continue  # 忽略该行可能出现的错误
    return -1  # 遍历10行后未找到


# --- 辅助函数：获取基准数据 (V4.8 - 自动查找表头) ---
def get_baseline_data(main_plan_path):
    log_messages = []
    default_date = pd.Timestamp('1900-01-01')
    existing_wo_set = set()

    try:
        log_messages.append(f"INFO: (baseline) 正在读取基准文件: {os.path.basename(main_plan_path)}")

        # --- [V4.8 修改] ---
        # 1. 第一次读取 (无表头)，用于查找表头
        try:
            df_no_header = pd.read_excel(main_plan_path, engine='openpyxl', header=None)
        except Exception as e:
            if "No such file" in str(e) or "FileNotFound" in str(e):
                raise FileNotFoundError  # 重新引发, 让外层try..except捕获
            log_messages.append(f"错误: (baseline) 第一次读取基准文件失败: {e}")
            raise e

        # 2. 查找表头行索引 (0-based)
        header_row_index = find_header_row_in_baseline(df_no_header)

        if header_row_index == -1:
            log_messages.append(
                f"警告: (baseline) 在基准文件的前10行中未找到包含 '工单单号', '工单单别' 和 'MOCTA' 的表头。")
            log_messages.append("警告: (baseline) 无法进行增量导入。将使用默认起始日期。")
            return default_date, existing_wo_set, log_messages

        log_messages.append(f"INFO: (baseline) 在基准文件的第 {header_row_index + 1} 行找到了表头。")

        # 3. 第二次读取 (使用找到的正确表头)
        # header=header_row_index (0-based)
        df = pd.read_excel(main_plan_path, engine='openpyxl', header=header_row_index)
        # --- [V4.8 修改结束] ---

        # --- [V4.7 逻辑] (现在应该可以正常工作了) ---
        date_col = None
        wo_num_col = None
        wo_type_col = None
        all_baseline_cols_cleaned = []  # 用于调试

        for col in df.columns:
            col_cleaned = str(col).strip()
            col_cleaned_upper = col_cleaned.upper()
            all_baseline_cols_cleaned.append(col_cleaned)

            if 'MOCTA' in col_cleaned_upper and 'CREATE' in col_cleaned_upper and 'DATE' in col_cleaned_upper:
                date_col = col
            if col_cleaned == '工单单号':
                wo_num_col = col
            if col_cleaned == '工单单别':
                wo_type_col = col

        if not date_col:
            log_messages.append(
                f"警告: (baseline) [调试] 基准文件中实际找到的(清理后)列名: {all_baseline_cols_cleaned}")
            log_messages.append("警告: (baseline) 找到了表头行，但未找到 'MOCTA' 日期列。将使用默认起始日期。")
            latest_date = default_date
        else:
            log_messages.append(f"INFO: (baseline) 找到基准日期列: '{date_col}'")
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            df_valid_dates = df.dropna(subset=[date_col])

            if df_valid_dates.empty:
                log_messages.append(f"INFO: (baseline) '{date_col}' 列中无有效日期。将使用默认起始日期。")
                latest_date = default_date
            else:
                latest_date = df_valid_dates[date_col].max()
                log_messages.append(f"INFO: (baseline) 找到最晚导入日期: {latest_date}")

        if wo_num_col and wo_type_col:
            df_keys = df.dropna(subset=[wo_num_col, wo_type_col])
            existing_wo_set = set(zip(
                df_keys[wo_num_col].astype(str).str.strip(),
                df_keys[wo_type_col].astype(str).str.strip()
            ))
            log_messages.append(f"INFO: (baseline) 基准文件中包含 {len(existing_wo_set)} 条已导入工单。")
        else:
            log_messages.append(f"警告: (baseline) 找到了表头行，但未找到 '工单单号' 或 '工单单别'。无法进行查重。")

        return latest_date, existing_wo_set, log_messages
        # --- [V4.7 逻辑结束] ---

    except FileNotFoundError:
        log_messages.append("INFO: (baseline) 未找到基准文件。将视为首次导入，导入所有工单。")
        return default_date, existing_wo_set, log_messages
    except Exception as e:
        log_messages.append(f"错误: (baseline) 读取基准文件失败 - {e}\n{traceback.format_exc()}")
        return default_date, existing_wo_set, log_messages


# --- [新增] 辅助函数：XLS 转换为 XLSX ---
def convert_xls_to_xlsx(xls_path):
    """
    使用 pywin32 (需要Excel) 将 .xls 文件转换为 .xlsx 文件。
    返回新的 .xlsx 文件路径。
    注意：这会覆盖同名的 .xlsx 文件（如果存在）。
    """
    pythoncom.CoInitialize()  # 确保COM库被正确初始化
    excel = None
    wb = None
    try:
        # 获取绝对路径
        xls_path_abs = os.path.abspath(xls_path)

        # 创建新的xlsx路径 (同名，不同后缀)
        new_xlsx_path = os.path.splitext(xls_path_abs)[0] + ".xlsx"

        # 启动Excel
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False  # 禁止弹出 "是否覆盖" 等提示

        # 打开 .xls 文件
        wb = excel.Workbooks.Open(xls_path_abs)

        # 另存为 .xlsx (FileFormat=51 是 xlOpenXMLWorkbook)
        # 如果文件已存在，将直接覆盖
        wb.SaveAs(new_xlsx_path, FileFormat=51)

        # 关闭工作簿
        wb.Close(False)  # False = 不保存更改 (因为已经SaveAs了)

        # 退出Excel
        excel.Quit()

        return new_xlsx_path

    except Exception as e:
        # 确保在出错时也尝试关闭Excel
        if wb:
            wb.Close(False)
        if excel:
            excel.Quit()
        # 抛出更详细的异常
        raise Exception(f"Excel COM 交互失败: {e}. 请确保已安装Excel且pywin32工作正常。")
    finally:
        # 清理COM对象
        del wb
        del excel
        pythoncom.CoUninitialize()


# --- 主要数据处理逻辑 (V4.6 - 模糊列查找) ---
def process_data(main_plan_path, shop_plan_path, schedule_template_path, output_dir):
    log_messages = []

    try:
        log_messages.append(
            f"INFO: 开始处理 V4.6 逻辑...\n基准文件: {main_plan_path}\n数据源: {shop_plan_path}\n模板: {schedule_template_path}")

        # --- 阶段一：确定需要导入的工单列表 ---

        latest_import_date, existing_wo_set, baseline_logs = get_baseline_data(main_plan_path)
        log_messages.extend(baseline_logs)
        latest_import_date_normalized = latest_import_date.normalize()
        log_messages.append(f"INFO: (baseline) 确定的基准日期为: {latest_import_date_normalized}")

        log_messages.append(f"INFO: 2. 正在读取 '分车间计划' (数据源): {os.path.basename(shop_plan_path)}")
        try:
            shop_plan_df = pd.read_excel(shop_plan_path, engine='openpyxl', dtype={
                '工单单号': str, '工单单别': str,
                '品号': str, '产品品号': str
            })

            # --- [V4.6 BUGFIX] ---
            # 迭代查找 date_col，绕过 'in' 检查的BUG
            date_col = None
            date_col_cleaned_name = None  # 存储清理后的名字 (MOCTA-CREATE-DATE)

            # 我们需要同时检查其他关键列
            required_shop_cols = ['工单单号', '工单单别', '品号', '产品品号', '预交货日']
            found_cols = set()

            all_cols_cleaned_for_debug = []
            rename_map = {}  # 用于重命名所有找到的列

            for col in shop_plan_df.columns:
                col_cleaned = str(col).strip()
                all_cols_cleaned_for_debug.append(col_cleaned)  # 收集所有清理后的列名
                col_cleaned_upper = col_cleaned.upper()  # 用于不区分大小写的比较

                # 检查必需的中文列
                if col_cleaned in required_shop_cols:
                    found_cols.add(col_cleaned)
                    # 如果原始列名带空格，也添加到重命名映射
                    if col != col_cleaned:
                        rename_map[col] = col_cleaned

                # 检查日期列 (模糊查找)
                if 'MOCTA' in col_cleaned_upper and 'CREATE' in col_cleaned_upper and 'DATE' in col_cleaned_upper:
                    date_col = col  # 存储原始列名 (可能带不可见字符)

                    # 确定它的“标准”名称
                    if 'MOCTA_CREATE_DATE' in col_cleaned_upper:
                        date_col_cleaned_name = 'MOCTA_CREATE_DATE'
                    else:
                        date_col_cleaned_name = 'MOCTA-CREATE-DATE'

                    # 添加到重命名映射
                    rename_map[col] = date_col_cleaned_name

            # --- [V4.6 BUGFIX 结束] ---

            # 检查所有必需的列
            missing = []
            for r_col in required_shop_cols:
                if r_col not in found_cols:
                    missing.append(r_col)

            if not date_col:
                missing.append('MOCTA-CREATE-DATE(或_DATE)')

            if missing:
                log_messages.append(f"错误: (shop_plan) 实际找到的(清理后的)列名: {all_cols_cleaned_for_debug}")
                raise ValueError(f"'分车间计划' 缺少关键列: {', '.join(missing)}")

            log_messages.append(
                f"INFO: (shop_plan) 找到数据源日期列: '{date_col}' (将重命名为: '{date_col_cleaned_name}')")

            # [V4.6 修改] 重命名所有关键列为“干净”的标准名称
            shop_plan_df.rename(columns=rename_map, inplace=True)
            log_messages.append(f"INFO: (shop_plan) 已自动清理关键数据列的列名。")

            # 现在我们可以100%安全地使用标准名称
            shop_plan_df['MOCTA-CREATE-DATE_dt'] = pd.to_datetime(shop_plan_df[date_col_cleaned_name], errors='coerce')
            shop_plan_df['MOCTA_DATE_ONLY'] = shop_plan_df['MOCTA-CREATE-DATE_dt'].dt.normalize()

            shop_plan_df['工单单号'] = shop_plan_df['工单单号'].astype(str).str.strip()
            shop_plan_df['工单单别'] = shop_plan_df['工单单别'].astype(str).str.strip()
            shop_plan_df['品号'] = shop_plan_df['品号'].astype(str).str.strip()
            shop_plan_df['产品品号'] = shop_plan_df['产品品号'].astype(str).str.strip()

        except Exception as e:
            if "File is not a zip file" in str(e):
                log_messages.append(
                    f"错误: 无法加载 '分车间计划'。文件 '{os.path.basename(shop_plan_path)}' 可能是 .xls 格式，但自动转换失败或未执行。")
            else:
                log_messages.append(f"错误: 无法加载 '分车间计划' 文件 - {e}\n{traceback.format_exc()}")
            return log_messages, False

        excluded_types = ['512', '522']
        original_count = len(shop_plan_df)
        shop_plan_df = shop_plan_df[~shop_plan_df['工单单别'].isin(excluded_types)]
        log_messages.append(f"INFO: (filter) 已排除 {original_count - len(shop_plan_df)} 条 '512'或'522' 单别的工单。")

        shop_plan_valid_df = shop_plan_df.dropna(subset=['MOCTA_DATE_ONLY'])
        on_date_df = shop_plan_valid_df[shop_plan_valid_df['MOCTA_DATE_ONLY'] == latest_import_date_normalized]
        after_date_df = shop_plan_valid_df[shop_plan_valid_df['MOCTA_DATE_ONLY'] > latest_import_date_normalized]
        log_messages.append(
            f"INFO: (filter) 找到 {len(on_date_df)} 条当天工单, {len(after_date_df)} 条未来工单 (已过滤)。")

        wo_to_import_set = set()
        on_date_wos = set(zip(on_date_df['工单单号'], on_date_df['工单单别']))
        new_on_date_wos = on_date_wos - existing_wo_set
        wo_to_import_set.update(new_on_date_wos)
        log_messages.append(f"INFO: (filter) 筛选出 {len(new_on_date_wos)} 条当天 *新* 工单。")

        after_date_wos = set(zip(after_date_df['工单单号'], after_date_df['工单单别']))
        new_after_date_wos = after_date_wos - existing_wo_set
        wo_to_import_set.update(new_after_date_wos)
        log_messages.append(f"INFO: (filter) 添加 {len(new_after_date_wos)} 条未来 *新* 工单。")

        if not wo_to_import_set:
            log_messages.append("INFO: 处理完成。没有需要导入的新工单。")
            return log_messages, True

        log_messages.append(f"INFO: (filter) 总计 {len(wo_to_import_set)} 条唯一工单 (单号, 单别) 待导入。")

        # --- 阶段二：获取输出数据 ---
        shop_plan_df['temp_wo_key'] = list(zip(shop_plan_df['工单单号'], shop_plan_df['工单单别']))
        data_to_export_df = shop_plan_df[shop_plan_df['temp_wo_key'].isin(wo_to_import_set)].copy()

        if data_to_export_df.empty:
            log_messages.append("警告: 找到了新工单，但在 '分车间计划' 中未找到任何匹配的行。")
            highlight_success, highlight_msg = highlight_rows_in_excel(shop_plan_path, wo_to_import_set)
            log_messages.append(highlight_msg)
            return log_messages, True

        log_messages.append(f"INFO: 3. 已从 '分车间计划' 提取 {len(data_to_export_df)} 条物料行用于处理。")

        # --- 阶段三：获取排序模板 ---
        log_messages.append(f"INFO: 4. 正在读取 '计划排产模板' (仅用于排序)...")
        try:
            template_df = pd.read_excel(schedule_template_path, engine='openpyxl', dtype={
                '产品编号': str,
                '料号': str
            })

            template_df.columns = template_df.columns.str.strip()

            required_tpl_cols = ['产品编号', '料号']
            if not all(col in template_df.columns for col in required_tpl_cols):
                missing = [col for col in required_tpl_cols if col not in template_df.columns]
                raise ValueError(f"'计划排产模板' 缺少排序关键列: {', '.join(missing)}")

            template_df['产品编号'] = template_df['产品编号'].astype(str).str.strip()
            template_df['料号'] = template_df['料号'].astype(str).str.strip()
            template_df['bom_sort_key'] = template_df.index
            bom_sort_map_df = template_df.drop_duplicates(subset=['产品编号', '料号'], keep='first')
            bom_sort_map = bom_sort_map_df.set_index(['产品编号', '料号'])['bom_sort_key'].to_dict()
            log_messages.append(f"INFO: 5. 已从模板创建 {len(bom_sort_map)} 条排序规则。")

        except Exception as e:
            log_messages.append(f"错误: 无法加载 '计划排产模板' 文件 - {e}\n{traceback.format_exc()}")
            return log_messages, False

        # --- 阶段四：合并排序键 ---
        log_messages.append("INFO: 6. 正在为数据应用BOM排序...")

        def get_bom_sort_key(row):
            key = (row['品号'], row['产品品号'])
            return bom_sort_map.get(key, float('inf'))

        data_to_export_df['bom_sort_key'] = data_to_export_df.apply(get_bom_sort_key, axis=1)

        unmapped_count = (data_to_export_df['bom_sort_key'] == float('inf')).sum()
        if unmapped_count > 0:
            log_messages.append(
                f"警告: {unmapped_count} 行数据在 '计划排产模板' 中未找到排序键，它们将被排在各自 '品号' 分组的末尾。")

        # --- 阶段五：格式化、输出、标黄 ---

        # 1. 格式化
        output_df = data_to_export_df

        log_messages.append("INFO: 7. 正在格式化输出列 (数据源: '分车间计划')...")

        output_df['预计产量_num'] = pd.to_numeric(output_df['预计产量'], errors='coerce').fillna(0)
        output_df['已生产量_num'] = pd.to_numeric(output_df['已生产量'], errors='coerce').fillna(0)
        output_df['未生产量'] = output_df['预计产量_num'] - output_df['已生产量_num']

        output_df.rename(columns={
            '备注': '订单PO#',
            '预计开工': '计划开工',
            '预计完工': '计划完工'
        }, inplace=True)

        # [V4.6 BUGFIX] 修复MOCTA-CREATE-DATE为空的问题
        # 我们现在使用 'date_col_cleaned_name' (MOCTA-CREATE-DATE 或 MOCTA_CREATE_DATE)
        if date_col_cleaned_name == 'MOCTA_CREATE_DATE':
            log_messages.append("INFO: (fix) 正在将 'MOCTA_CREATE_DATE' 数据复制到 'MOCTA-CREATE-DATE' 列。")
            output_df['MOCTA-CREATE-DATE'] = output_df['MOCTA_CREATE_DATE']

        # 如果 'date_col_cleaned_name' 已经是 'MOCTA-CREATE-DATE'，则无需操作

        final_column_order = [
            '工单单号', '工单单别', '工作中心名称', '产品品号', '产品品名', '产品规格', '订单PO#',
            '单位', '预计产量', '已生产量', '未生产量', '标准工时', '总工时', '开工日期',
            '完工日期', '计划开工', '计划完工', '预交货日', '品号', '品名', '订单',
            '开单日期', '客户单号', '订单日期', '订单数量', '齐套率', '用户自定义字段7',
            '规格', '简称', '[品号分类三]', '审核码', 'MOCTA-CREATE-DATE', '要求物料到位时间'
        ]

        for col in final_column_order:
            if col not in output_df.columns:
                output_df[col] = pd.NA

                # 2. 排序 (V4.2 逻辑)
        if '订单PO#' not in output_df.columns: output_df['订单PO#'] = pd.NA
        output_df['订单PO#'] = output_df['订单PO#'].fillna('').astype(str)

        output_df['预交货日'] = pd.to_datetime(output_df['预交货日'], errors='coerce')

        log_messages.append("INFO: 8. 正在按 '订单PO#' -> '预交货日' -> '品号' -> 'BOM模板顺序' 排序...")
        output_df.sort_values(
            by=['订单PO#', '预交货日', '品号', 'bom_sort_key'],
            ascending=[True, True, True, True],
            inplace=True
        )

        # 3. 清理并重排
        helper_cols_to_drop = [
            'bom_sort_key', 'temp_wo_key', 'MOCTA-CREATE-DATE_dt', 'MOCTA_DATE_ONLY',
            '预计产量_num', '已生产量_num'
        ]
        # 如果原始日期列被重命名了，我们也删掉它
        if date_col_cleaned_name == 'MOCTA_CREATE_DATE':
            helper_cols_to_drop.append('MOCTA_CREATE_DATE')

        output_df = output_df.drop(columns=helper_cols_to_drop, errors='ignore')

        output_df = output_df.reindex(columns=final_column_order)

        # 4. 保存
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"工单工序计划表_{timestamp}.xlsx"
        output_file_path = os.path.join(output_dir, output_filename)

        try:
            output_df.to_excel(output_file_path, index=False, engine='openpyxl')
            log_messages.append(f"成功: 9. 成功生成输出文件: {output_file_path} (共 {len(output_df)} 行)")
        except Exception as e:
            log_messages.append(f"错误: 保存输出文件失败 - {e}")
            return log_messages, False

        # 5. 标黄
        log_messages.append(f"INFO: 10. 正在更新源文件 '{os.path.basename(shop_plan_path)}' (标黄)...")
        highlight_success, highlight_msg = highlight_rows_in_excel(shop_plan_path, wo_to_import_set)
        log_messages.append(highlight_msg)

        if not highlight_success:
            log_messages.append("警告: 数据已导出，但在标黄 '分车间计划' 文件时发生错误。请检查上述日志。")

        log_messages.append("\nINFO: 全部处理完成!")
        return log_messages, True

    except Exception as e:
        log_messages.append(f"处理过程中发生意外全局错误: {e}\n{traceback.format_exc()}")
        return log_messages, False



# --- GUI 代码 ---
main_plan_path_var = ""
shop_plan_path_var = ""
schedule_template_path_var = ""
output_directory = os.path.join(os.path.expanduser("~"), "Desktop")
if not os.path.exists(output_directory):
    output_directory = os.getcwd()


# (GUI选择文件的函数 select_... 无修改)
def select_main_plan_file():
    global main_plan_path_var
    filepath = filedialog.askopenfilename(
        title="1. 选择 主生产计划 (基准文件)",
        filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
    )
    if filepath:
        main_plan_path_var = filepath
        main_plan_path_label.config(text=os.path.basename(filepath))
        log_text.insert(tk.END, f"已选择基准文件: {filepath}\n")


def select_shop_plan_file():
    global shop_plan_path_var
    filepath = filedialog.askopenfilename(
        title="2. 选择 分车间计划 (数据源,将被标黄)",
        filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))  # 允许选择 .xls
    )
    if filepath:
        shop_plan_path_var = filepath
        shop_plan_path_label.config(text=os.path.basename(filepath))
        log_text.insert(tk.END, f"已选择数据源文件: {filepath}\n")


def select_template_file():
    global schedule_template_path_var
    filepath = filedialog.askopenfilename(
        title="3. 选择 计划排产模板 (明细)",
        filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
    )
    if filepath:
        schedule_template_path_var = filepath
        template_path_label.config(text=os.path.basename(filepath))
        log_text.insert(tk.END, f"已选择明细模板文件: {filepath}\n")


def select_output_directory():
    global output_directory
    dirpath = filedialog.askdirectory(title="选择输出目录")
    if dirpath:
        output_directory = dirpath
        output_path_display.config(text=output_directory)
        log_text.insert(tk.END, f"已选择输出目录: {output_directory}\n")
        update_initial_log_message()


# --- [修改] run_processing 函数 ---
def run_processing():
    global output_directory
    global shop_plan_path_var  # <-- [修改] 声明为 global, 因为可能需要更新它

    # 文件检查（允许基准文件为空）
    if not shop_plan_path_var:
        messagebox.showerror("错误", "请先选择 '分车间计划' (数据源)！");
        return
    if not schedule_template_path_var:
        messagebox.showerror("错误", "请先选择 '计划排产模板' (明细)！");
        return

    if not main_plan_path_var:
        log_text.insert(tk.END, "警告: 未选择 '主生产计划' (基准文件)。\n")
        log_text.insert(tk.END, "将按首次运行处理，导入 '分车间计划' 中的所有工单。\n")
        root.update_idletasks()

    # 清理旧日志 (移到前面)
    current_log = log_text.get("1.0", tk.END)
    separator = "-------------------------------------\n"
    separator_index = current_log.find(separator)
    if separator_index != -1:
        delete_start_char_index = separator_index + len(separator)
        line_col_separator_end = log_text.index(f"1.0 + {delete_start_char_index} chars")
        log_text.delete(line_col_separator_end, tk.END)
    else:
        log_text.delete(1.0, tk.END)
        update_initial_log_message()

    run_button.config(state=tk.DISABLED)

    # --- [新增] XLS to XLSX 转换逻辑 ---
    if shop_plan_path_var.lower().endswith(".xls"):
        log_text.insert(tk.END, f"INFO: 检测到 '.xls' 文件: {os.path.basename(shop_plan_path_var)}\n")
        log_text.insert(tk.END,
                        "INFO: 标黄功能(openpyxl)无法处理 .xls。\nINFO: 正在尝试使用 Excel 自动转换为 .xlsx...\n")
        log_text.insert(tk.END, "INFO: 这需要本地安装了 Microsoft Excel，请稍候...\n")
        root.update_idletasks()

        try:
            # 调用转换函数
            new_xlsx_path = convert_xls_to_xlsx(shop_plan_path_var)

            log_text.insert(tk.END, f"INFO: 转换成功。新文件已创建: {os.path.basename(new_xlsx_path)}\n")
            log_text.insert(tk.END, f"INFO: 后续操作 (包括标黄) 将针对此 .xlsx 文件进行。\n")

            # [关键] 更新全局路径变量和GUI标签，指向新生成的 .xlsx 文件
            shop_plan_path_var = new_xlsx_path
            shop_plan_path_label.config(text=os.path.basename(shop_plan_path_var))

        except Exception as e:
            log_text.insert(tk.END, f"错误: 自动转换 .xls 文件失败: {e}\n")
            log_text.insert(tk.END, "请手动将 '分车间计划' 文件另存为 .xlsx 格式，然后重新运行程序。\n")
            messagebox.showerror("转换失败",
                                 f"无法将 .xls 转换为 .xlsx: {e}\n\n请手动将 '分车间计划' 另存为 .xlsx 格式后重试。")
            run_button.config(state=tk.NORMAL)
            return
    # --- [新增结束] ---

    if not os.path.exists(output_directory):
        try:
            os.makedirs(output_directory, exist_ok=True)
            log_text.insert(tk.END, f"INFO: 输出目录已创建: {output_directory}\n")
        except Exception as e:
            messagebox.showerror("错误", f"无法创建输出目录 {output_directory}: {e}. 请重新选择或检查权限。")
            run_button.config(state=tk.NORMAL)  # [修改] 确保按钮在出错时恢复
            return

    output_path_display.config(text=f"{output_directory}")

    log_text.insert(tk.END, "\nINFO: 开始处理，请稍候...\n")
    log_text.insert(tk.END,
                    f"INFO: 请确保 '分车间计划' ({os.path.basename(shop_plan_path_var)}) 在此期间已关闭，以便程序能标黄它。\n")
    root.update_idletasks()

    # (调用 process_data 的逻辑无修改)
    results, success = process_data(main_plan_path_var, shop_plan_path_var, schedule_template_path_var,
                                    output_directory)

    for msg in results: log_text.insert(tk.END, msg + "\n")
    log_text.see(tk.END)

    # (显示结果的逻辑无修改)
    log_summary = "".join(results)
    if success:
        if "没有需要导入的新工单" in log_summary:
            messagebox.showinfo("完成", "处理完成，没有找到需要导入的新工单。")
        elif "成功: (highlight)" in log_summary and "成功: 7." in log_summary:
            messagebox.showinfo("完成",
                                f"数据处理已完成！新文件已生成，且 '分车间计划' ({os.path.basename(shop_plan_path_var)}) 已标黄。\n输出目录: {output_directory}。")
        elif "警告: 数据已导出，但在标黄" in log_summary:
            messagebox.showwarning("处理部分完成",
                                   f"数据处理流程已完成，新文件已生成，但标黄 '分车间计划' 时发生错误。请仔细检查日志！")
        else:
            messagebox.showinfo("完成",
                                f"数据处理已完成。输出文件已生成在: {output_directory}。请检查日志确认详情。")
    else:
        messagebox.showerror("处理异常", "数据处理过程中发生错误，操作可能未完成。详情请查看日志。")

    run_button.config(state=tk.NORMAL)


# (GUI 布局和 mainloop 代码无修改)
def update_initial_log_message():
    log_text.delete(1.0, tk.END)
    initial_log_content = initial_log_message_template.format(
        output_directory=output_directory
    )
    log_text.insert(tk.END, initial_log_content)


root = tk.Tk()
root.title("增量工单导入工具 (v2.1 - 支持xls转换)")
root.geometry("750x800")

file_frame = tk.Frame(root, padx=10, pady=10);
file_frame.pack(fill=tk.X)

tk.Label(file_frame, text="1. 主生产计划 (基准):", width=20, anchor='w').grid(row=0, column=0, padx=5, pady=5,
                                                                              sticky='w')
main_plan_path_label = tk.Label(file_frame, text="[可选] (若不选,则导入全部)", width=45, anchor='w', relief="sunken",
                                borderwidth=1);
main_plan_path_label.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
tk.Button(file_frame, text="选择文件...", command=select_main_plan_file, width=15).grid(row=0, column=2, padx=5, pady=5)

tk.Label(file_frame, text="2. 分车间计划 (数据源):", width=20, anchor='w').grid(row=1, column=0, padx=5, pady=5,
                                                                                sticky='w')
shop_plan_path_label = tk.Label(file_frame, text="[必选] (此文件将被标黄)", width=45, anchor='w', relief="sunken",
                                borderwidth=1);
shop_plan_path_label.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
tk.Button(file_frame, text="选择文件...", command=select_shop_plan_file, width=15).grid(row=1, column=2, padx=5, pady=5)

tk.Label(file_frame, text="3. 计划排产模板 (明细):", width=20, anchor='w').grid(row=2, column=0, padx=5, pady=5,
                                                                                sticky='w')
template_path_label = tk.Label(file_frame, text="[必选]", width=45, anchor='w', relief="sunken", borderwidth=1);
template_path_label.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
tk.Button(file_frame, text="选择文件...", command=select_template_file, width=15).grid(row=2, column=2, padx=5, pady=5)

tk.Label(file_frame, text="输出目录:", width=20, anchor='w').grid(row=3, column=0, padx=5, pady=5, sticky='w')
output_path_display = tk.Label(file_frame, text=f"{output_directory}", width=45, anchor='w', relief="sunken",
                               borderwidth=1);
output_path_display.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
tk.Button(file_frame, text="选择目录...", command=select_output_directory, width=15).grid(row=3, column=2, padx=5,
                                                                                          pady=5)

file_frame.columnconfigure(1, weight=1)

run_button = tk.Button(root, text="开始导入数据", command=run_processing, font=('Arial', 12, 'bold'), bg='lightblue',
                       width=20, pady=10);
run_button.pack(pady=15)

log_label = tk.Label(root, text="处理日志:", font=('Arial', 10, 'bold'), anchor='w');
log_label.pack(fill=tk.X, padx=10, pady=(0, 5))

log_text = scrolledtext.ScrolledText(root, height=25, wrap=tk.WORD, relief="solid", borderwidth=1);
log_text.pack(padx=10, pady=(0, 10), fill=tk.BOTH, expand=True)

initial_log_message_template = (
    "程序已启动。请按顺序选择文件：\n"
    "1. [主生产计划] (可选): 已导入的数据基准。\n"
    "   -> 如果不选，将导入'分车间计划'中的所有工单。\n"
    "2. [分车间计划] (必选): 包含新工单的数据源。\n"
    "   -> 此文件将被修改 (新导入行会标黄)。\n"
    "   -> (v2.1) 支持 .xls 文件自动转换 (需安装Excel)。\n"
    "3. [计划排产模板] (必选): 包含工单物料明细。\n\n"
    "当前输出目录: {output_directory}\n"
    "-------------------------------------\n"
)

update_initial_log_message()

root.mainloop()
